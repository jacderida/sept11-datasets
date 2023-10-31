#!/usr/bin/env python

import os
import sys
import re
import subprocess
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm


def get_file_description(filepath):
    cmd = ["file", "-b", filepath]
    result = subprocess.run(cmd, capture_output=True, text=True)
    return result.stdout.strip()


def categorize_file(file_type):
    if "video" in file_type or "AVI" in file_type:
        return "Video"
    elif "image" in file_type or "JPEG" in file_type:
        matches = re.findall(r"(\d+)x(\d+)", file_type)
        if matches:
            width, height = map(int, matches[-1])
            resolution = width * height

            if resolution < 800 * 600:
                return "Low-res Image"
            elif resolution < 1920 * 1080:
                return "Medium-res Image"
            else:
                return "High-res Image"
        else:
            return "Image (Resolution Unknown)"
    elif "PDF document" in file_type:
        return "PDF"
    elif "ASCII text" in file_type or "text" in file_type:
        return "Text"
    else:
        return "Other"


def adjust_column_widths(headers, sheet):
    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        max_length = 0
        column = sheet[col_letter]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[col_letter].width = adjusted_width


def generate_report(file_path, file_descriptions):
    workbook = Workbook()
    sheet = workbook.create_sheet("Release Files", 0)
    headers = ["Path", "Category", "Description"]
    for col_num, header in enumerate(headers, 1):
        col_letter = chr(64 + col_num)
        sheet[f"{col_letter}1"] = header
    row_num = 2
    for description in file_descriptions:
        sheet[f"A{row_num}"] = description[0]
        sheet[f"B{row_num}"] = description[1]
        sheet[f"C{row_num}"] = description[2]
        row_num += 1
    adjust_column_widths(headers, sheet)
    sheet.freeze_panes = "A2"
    workbook.save(file_path)


def main(dir_path):
    file_descriptions = []
    categories = set()

    total_files = sum(len(files) for _, _, files in os.walk(dir_path))

    with tqdm(total=total_files, desc="Processing files") as pbar:
        for root, _, files in os.walk(dir_path):
            for file in files:
                file_path = os.path.join(root, file)
                file_description = get_file_description(file_path)
                file_category = categorize_file(file_description)
                categories.add(file_category)
                file_descriptions.append(
                    (
                        file_path,
                        file_category,
                        file_description
                    )
                )
                pbar.update(1)

    print("Summary:")
    for category in categories:
        items = [x for x in file_descriptions if x[1] == category]
        print(f"{category}: {len(items)}")

    output_filename = f"{os.path.basename(dir_path)}.xlsx"
    print(f"Saving {output_filename}")
    if os.path.exists(output_filename):
        os.remove(output_filename)
    generate_report(output_filename, file_descriptions)


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: script.py <directory_path>")
        sys.exit(1)
    dir_path = sys.argv[1]
    main(dir_path)
