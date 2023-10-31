#!/usr/bin/env python
import os
import sqlite3
import sys
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass
class ReleaseData:
    id: str
    date: str
    name: str
    directory: str
    file_count: int
    notes: str
    size: int
    torrent_url: str
    download_url: str
    verification_outcome: str

    @staticmethod
    def from_row(row):
        id = row[0]
        date = row[1]
        name = row[2]
        directory = row[3]
        file_count = row[4]
        notes = row[5]
        size = row[6]
        torrent_url = row[7]
        download_url = row[8]
        verification_outcome = row[9]
        return ReleaseData(
            id, date, name,
            directory, file_count, notes, size, torrent_url, download_url, verification_outcome)


@dataclass
class IncompleteReleaseData:
    name: str
    file_count: int
    corrupt_file_count: int
    missing_file_count: int
    size: int
    corrupt_size: int
    missing_size: int
    notes: str


@dataclass
class IncompleteFilesData:
    release_name: str
    file_path: str
    size: int
    status: str


@dataclass
class Release14LinksData:
    path: str
    url: str


def get_db_path():
    if "HOME" in os.environ:
        app_data_path = os.path.join(os.environ["HOME"], ".local", "share", "sept11-datasets")
    else:
        raise Exception("Could not find home directory")
    if not os.path.exists(app_data_path):
        os.makedirs(app_data_path)
    database_path = os.path.join(app_data_path, "releases.db")
    return database_path


def get_releases():
    with sqlite3.connect(get_db_path()) as conn:
        cursor = conn.cursor()
        cursor.execute("""
        SELECT
            id, date, name,
            directory, file_count, notes, size, torrent_url, download_url, verification_outcome
        FROM releases
        """)
        releases = []
        rows = cursor.fetchall()
        for row in rows:
            release = ReleaseData.from_row(row)
            releases.append(release)
        return releases


def get_incomplete_release_data():
    with sqlite3.connect(get_db_path()) as conn:
        incomplete_release_data = []
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, name, file_count, notes, size FROM releases
            WHERE verification_outcome = 'INCOMPLETE'
        """)
        release_rows = cursor.fetchall()
        for row in release_rows:
            id = row[0]
            name = row[1]
            file_count = row[2]
            notes = row[3]
            size = row[4]

            cursor.execute(
                """
                SELECT COUNT(*) FROM incomplete_files WHERE release_id = ? AND status = 'CORRUPTED'
                """, (id,))
            corrupt_file_count = cursor.fetchone()[0]
            cursor.execute(
                """
                SELECT COUNT(*) FROM incomplete_files WHERE release_id = ? AND status = 'MISSING'
                """, (id,))
            missing_file_count = cursor.fetchone()[0]

            cursor.execute(
                """
                SELECT SUM(size) FROM incomplete_files WHERE release_id = ? AND status = 'CORRUPTED'
                """, (id,))
            corrupt_size = cursor.fetchone()[0]
            cursor.execute(
                """
                SELECT SUM(size) FROM incomplete_files WHERE release_id = ? AND status = 'MISSING'
                """, (id,))
            missing_size = cursor.fetchone()[0]
            if not missing_size:
                missing_size = 0

            incomplete_release_data.append(IncompleteReleaseData(
                name, file_count, corrupt_file_count, missing_file_count,
                size, corrupt_size, missing_size, notes
            ))
        return incomplete_release_data


def get_incomplete_files_data():
    with sqlite3.connect(get_db_path()) as conn:
        incomplete_files_data = []
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, name FROM releases WHERE verification_outcome = 'INCOMPLETE'
        """)
        release_rows = cursor.fetchall()
        for row in release_rows:
            id = row[0]
            name = row[1]
            cursor.execute(
                """
                SELECT file_path, size, status FROM incomplete_files WHERE release_id = ?
                """, (id,))
            incomplete_rows = cursor.fetchall()
            for incomplete_row in incomplete_rows:
                incomplete_files_data.append(
                    IncompleteFilesData(
                        name, incomplete_row[0], incomplete_row[1], incomplete_row[2]))
        return incomplete_files_data


def get_release_14_links_data():
    with sqlite3.connect(get_db_path()) as conn:
        release_14_links_data = []
        cursor = conn.cursor()
        cursor.execute("""
            SELECT directory_path, base_url FROM release_14_links
        """)
        rows = cursor.fetchall()
        for row in rows:
            path = row[0]
            url = row[1]
            release_14_links_data.append(Release14LinksData(path, url))
        return release_14_links_data


def bytes_to_human_readable(bytes):
    if bytes is None:
        return "N/A"

    TB = 1024 ** 4
    GB = 1024 ** 3
    MB = 1024 ** 2
    KB = 1024

    if bytes == 0:
        return "0"
    elif bytes >= TB:
        return "{:.2f} TB".format(bytes / TB)
    elif bytes >= GB:
        return "{:.2f} GB".format(bytes / GB)
    elif bytes >= MB:
        return "{:.2f} MB".format(bytes / MB)
    elif bytes >= KB:
        return "{:.2f} KB".format(bytes / KB)
    else:
        return "{} bytes".format(bytes)


def adjust_column_widths(headers, sheet):
    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        max_length = 0
        column = sheet[col_letter]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[col_letter].width = adjusted_width


def build_releases_sheet(workbook, releases):
    lime_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    cyan_fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")

    print("Creating releases sheet")
    sheet = workbook.create_sheet("Releases", 0)
    headers = ["Date", "Name", "Files", "Size", "Status", "Download Link"]
    for col_num, header in enumerate(headers, 1):
        col_letter = chr(64 + col_num)
        sheet[f"{col_letter}1"] = header
    row_num = 2
    for release in releases:
        sheet[f"A{row_num}"] = release.date
        sheet[f"B{row_num}"] = release.name
        sheet[f"C{row_num}"] = release.file_count if release.file_count else "N/A"
        sheet[f"D{row_num}"] = bytes_to_human_readable(release.size)
        sheet[f"E{row_num}"] = release.verification_outcome
        if release.download_url:
            cell = sheet[f"F{row_num}"]
            cell.value = release.download_url
            cell.hyperlink = release.download_url
            cell.style = "Hyperlink"
        else:
            sheet[f"F{row_num}"] = "N/A"
        
        if release.verification_outcome == "VERIFIED":
            sheet.cell(row=row_num, column=5).fill = lime_fill
        elif release.verification_outcome == "MISSING":
            sheet.cell(row=row_num, column=5).fill = red_fill
        elif release.verification_outcome == "NO TORRENT":
            sheet.cell(row=row_num, column=5).fill = orange_fill
        elif release.verification_outcome == "INCOMPLETE":
            sheet.cell(row=row_num, column=5).fill = cyan_fill
        row_num += 1

    adjust_column_widths(headers, sheet)

    # Left justify the 'Files' column
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal="left")

    # Make headers and status column text bold
    bold_font = Font(name="Calibri", bold=True)
    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"].font = bold_font
    for cell in sheet["E"]:
        cell.font = bold_font
        cell.alignment = Alignment(horizontal="center")

    sheet.freeze_panes = "A2"
    return sheet


def build_incomplete_releases_sheet(workbook, incomplete_release_data):
    print("Creating incomplete releases sheet")
    sheet = workbook.create_sheet("Incomplete Releases", 1)
    headers = [
        "Name", "Total Files",
        "Corrupt Files", "Missing Files", "Total Size",
        "Corrupt Size", "Missing Size", "Notes"
    ]
    for col_num, header in enumerate(headers, 1):
        col_letter = chr(64 + col_num)
        sheet[f"{col_letter}1"] = header
    row_num = 2
    for release in incomplete_release_data:
        sheet[f"A{row_num}"] = release.name
        sheet[f"B{row_num}"] = release.file_count
        sheet[f"C{row_num}"] = release.corrupt_file_count
        sheet[f"D{row_num}"] = release.missing_file_count
        sheet[f"E{row_num}"] = bytes_to_human_readable(release.size)
        sheet[f"F{row_num}"] = bytes_to_human_readable(release.corrupt_size)
        sheet[f"G{row_num}"] = bytes_to_human_readable(release.missing_size)
        sheet[f"H{row_num}"] = release.notes
        row_num += 1

    # Left justify a few columns
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=2, max_col=4):
        for cell in row:
            cell.alignment = Alignment(horizontal="left")

    # Make headers text bold
    bold_font = Font(name="Calibri", bold=True)
    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"].font = bold_font

    adjust_column_widths(headers, sheet)


def build_incomplete_files_sheet(workbook, incomplete_files_data):
    print("Creating incomplete files sheet")
    sheet = workbook.create_sheet("Corrupt or Missing Files", 2)
    headers = ["Release", "Size", "Status", "Path"]
    for col_num, header in enumerate(headers, 1):
        col_letter = chr(64 + col_num)
        sheet[f"{col_letter}1"] = header
    row_num = 2
    for file in incomplete_files_data:
        sheet[f"A{row_num}"] = file.release_name
        sheet[f"B{row_num}"] = bytes_to_human_readable(file.size)
        sheet[f"C{row_num}"] = file.status
        sheet[f"D{row_num}"] = file.file_path
        row_num += 1

    # Left justify a few columns
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=2, max_col=3):
        for cell in row:
            cell.alignment = Alignment(horizontal="left")

    # Make headers text bold
    bold_font = Font(name="Calibri", bold=True)
    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"].font = bold_font

    adjust_column_widths(headers, sheet)
    sheet.freeze_panes = "A2"


def build_release14_links_sheet(workbook, release_14_links_data):
    print("Creating incomplete files sheet")
    sheet = workbook.create_sheet("Release 14 Links", 3)
    headers = ["Collection", "Link"]
    for col_num, header in enumerate(headers, 1):
        col_letter = chr(64 + col_num)
        sheet[f"{col_letter}1"] = header
    row_num = 2
    for item in release_14_links_data:
        sheet[f"A{row_num}"] = item.path
        cell = sheet[f"B{row_num}"]
        cell.value = item.url
        cell.hyperlink = item.url
        cell.style = "Hyperlink"
        row_num += 1

    # Make headers text bold
    bold_font = Font(name="Calibri", bold=True)
    for col_num, _ in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        sheet[f"{col_letter}1"].font = bold_font

    adjust_column_widths(headers, sheet)
    sheet.freeze_panes = "A2"


def generate_report(file_path):
    releases = get_releases()
    incomplete_release_data = get_incomplete_release_data()
    incomplete_files_data = get_incomplete_files_data()
    release_14_links_data = get_release_14_links_data()
    release_14_links_data.sort(key=lambda x: x.path)

    workbook = Workbook()
    build_releases_sheet(workbook, releases)
    build_incomplete_releases_sheet(workbook, incomplete_release_data)
    build_incomplete_files_sheet(workbook, incomplete_files_data)
    build_release14_links_sheet(workbook, release_14_links_data)
    workbook.save(file_path)


def main():
    if os.path.exists("report.xlsx"):
        os.remove("report.xlsx")
    generate_report("report.xlsx")
    return 0


if __name__ == "__main__":
    sys.exit(main())
